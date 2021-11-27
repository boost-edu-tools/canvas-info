# Copyright 2021 Huub de Beer <h.t.d.beer@tue.nl>
#
# Licensed under the EUPL, Version 1.2 or later. You may not use this work
# except in compliance with the EUPL. You may obtain a copy of the EUPL at:
#
# https://joinup.ec.europa.eu/software/page/eupl
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the EUPL is distributed on an "AS IS" basis, WITHOUT
# WARRANTY OR CONDITIONS OF ANY KIND, either express or implied. See the EUPL
# for the specific language governing permissions and limitations under the
# licence.
"""Map student IDs from Canvas to Git and vice versa.

To map IDs from Git to Canvas, and the other way around, use the CanvasGitMap
class. The CanvasGitMap class is a table with, for each student in a course,
at least two columns: canvas_id and git_id. Furthermore, for readability, you
can have more columns in the table. For example, an email address or student
name is convenient.

A CanvasGitMap is valid if each student has both a Canvas ID and a Git ID, and
both IDs are unique in the map.

The CanvasGitMap extends the Table class. The Table class offers basic
functionality to deal with a table, like access to its column names, its rows,
and reading from and writing to CSV files.
"""
import csv
from pathlib            import Path
from typing             import List

from .canvas_api.course import Course
from .canvas_api.user   import PUBLIC_USER_FIELDS
from .common            import warn, inform
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet import table

CANVAS_ID               = "canvas_id"
CANVAS_LOGIN_ID         = "login_id"
FIELD_SEP               = ","
GIT_ID                  = "GitID"
SHORT_NAME              = 'short_name'
FULL_NAME               = 'FullName'
GROUP                   = 'Group'
ID                      = 'ID'
EMAIL                   = 'Mail'
NAME                    = 'Name'
HEAD                    = 5

class Table:
    """Table"""

    def __init__(self, data : List):
        self._data = data

    @classmethod
    def load(cls, path : Path):
        """Load Table from a csv file."""
        with path.open() as csv_file:
            return cls(csv.DictReader(csv_file, delimiter = FIELD_SEP))

    @classmethod
    def loadExcel(cls, path : str):
        """ Get all tables from a given workbook. Returns a dictionary of tables.
        Requires a filename, which includes the file path and filename. """

        # Load the workbook, from the filename, setting read_only to False
        wb = load_workbook(filename=path, keep_vba=False, data_only=True, keep_links=False)

        # Get a list of all rows
        rows_list = []

        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]

            # Get each table in the worksheet
            for tbl in ws.tables.values():
                # Grab the 'data' from the table
                data = ws[tbl.ref]

                #get the first header row
                keys = []
                for k in data[0]:
                    keys.append(k.value)

                for row in data[1:]:
                    # Get a list of all columns in each row
                    cols = {}
                    for i, col in enumerate(row):
                        cols[keys[i]] = col.value
                    rows_list.append(cols)

        return cls(rows_list)

    def write(self, path : Path):
        # """Write this Canvas-Git map to csv file."""
        with path.open("w", encoding='utf-8-sig', newline='') as csv_file:

            csv_writer = csv.DictWriter(
                    csv_file,
                    delimiter   = FIELD_SEP,
                    fieldnames  = list(self.columns()),
                    )

            csv_writer.writeheader()

            for row in self.rows():
                csv_writer.writerow(row)

    def writeExcel(self, path: str):
        wb = Workbook()
        ws = wb.active

        columns = list(self.columns())
        ws.append(columns)

        for row in self.rows():
            ws.append(list(row.values()))

        tab = table.Table(displayName="table1", ref='A1:F' + str(len(self._data)+1))

        ws.column_dimensions["B"].width = 15 #name column
        ws.column_dimensions["C"].width = 25 #full name column
        ws.column_dimensions["F"].width = 45 #emil column

        # Add a default style with striped rows and banded columns
        style = table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        '''
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unque through out defined names and all other table name.
        '''
        ws.add_table(tab)
        wb.save(path)

    def columns(self):
        """Generator for the column names of this Table."""
        columns = []

        if len(self._data) > 0:
            columns = list(self._data[0].keys())

        return columns

    def rows(self):
        """Generator for each row of this Table."""
        for row in self._data:
            yield row

    def empty(self):
        """Return true if this table is empty, false otherwise."""
        return 0 == len(self._data)

    def get_stu_info(self) -> list:
        student_info = []
        for row in self.rows():
            student_info.append({"group":row[GROUP], "email2git": {row[EMAIL][:-15]:row[GIT_ID]}})

        return student_info

class CanvasGitMap(Table):
    """Map Canvas IDs to Git IDs and vice versa. The CanvasGitMap uses a
    data table with at least two columns, "git_id" and "canvas_id", to perform
    the mapping."""

    """Added: Map Canvas IDs to Email (column "email") for student YAML file"""

    def __init__(self, data : List):
        super().__init__(data)

        self._canvas2git = {}
        self._git2canvas = {}

        for row in self.rows():
            canvas_id   = row[CANVAS_ID]
            git_id      = row[GIT_ID]

            _check_id("Canvas", canvas_id, self._canvas2git)
            _check_id("Git", git_id, self._git2canvas)

            self._canvas2git[canvas_id] = row[GIT_ID]
            self._git2canvas[git_id]    = row[CANVAS_ID]

    def canvas2git(self, canvas_id : str) -> str:
        """Convert a Canvas ID to the correspondibg Git ID."""
        if canvas_id in self._canvas2git:
            return self._canvas2git[canvas_id]

        raise ValueError(f"Canvas ID '{canvas_id}' not mapped to a Git ID.")

    def git2canvas(self, git_id : str) -> str:
        """Convert a Git ID to the corresponding Canvas ID."""
        if git_id in self._git2canvas:
            return self._git2canvas[git_id]

        raise ValueError(f"Git ID '{git_id}' not mapped to a Canvas ID.")

# Guide the user in creating a potential Canvas-Git mapping table for a
# Canvas course.

ASK_GIT_ID          = ("Which column do you want to use as the students' "
                        "Git ID in the Canvas-Git mapping table?")
ASK_EXTRA_COLUMNS   = ("Which extra columns to you want to add to the "
                        "Canvas-Git mapping table? "
                        "Press SPACE to select an item; multiple items "
                        "can be selected. Press ENTER to confirm your choice.")

def canvas_git_map_table_wizard(course : Course, group_category : str = None) -> Table:
    """Create a Canvas-Git map CSV file."""
    inform("Getting the students' infomation...")
    students = course.students()

    if len(students) <= 0:
        warn((f"No students found for course '{course.name}'."))
        return Table([])

    inform((f"Found {len(students)} students for this course. "))

    inform("Getting the information of groups...")
    group_members = course.group_members(group_category)

    canvas_id_key = CANVAS_LOGIN_ID

    git_id_key = 'sis_user_id'
    email_key = 'email'
    student_id = 'id'

    data = []

    for student in students:
        row = {}
        fields = student.fields()

        if student_id in fields:
            user_id = fields[student_id]
            if user_id in group_members:
                row[GROUP] = group_members[user_id]
            else:
                row[GROUP] = ""
        else:
            row[GROUP] = ""

        email = ""
        if email_key in fields:
            email = fields[email_key]
            row[NAME] = email[:-15].split(".")[-1]
        else:
            row[NAME] = ""

        if SHORT_NAME in fields:
            row[FULL_NAME] = fields[SHORT_NAME]
        else:
            row[FULL_NAME] = ""

        if canvas_id_key in fields:
            row[ID] = fields[canvas_id_key]
        else:
            row[ID] = ""

        if git_id_key in fields:
            row[GIT_ID] = fields[git_id_key]
        else:
            row[GIT_ID] = ""

        row[EMAIL] = email

        data.append(row)

    data = sorted(data, key=lambda d: d[GROUP])

    return Table(data)

# Private functions
def _check_id(service : str, service_id : str, service_map : str) -> bool:
    if not service_id:
        raise ValueError(f"The {service} ID cannot be empty.")

    if service_id in service_map:
        raise ValueError(f"The {service} ID '{service_id}' is not unique.")
